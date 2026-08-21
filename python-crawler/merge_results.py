"""
merge_results.py
================
Собирает все файлы из папки INPUT_DIR, извлекает записи со score >= 8,
дедублицирует по домену, объединяет в один Excel файл.

Поддерживаемые форматы:
- batch_*.xlsx (листы Results / High Score)
- validated_vendors.xlsx
- classified_sites_*.csv / results*.csv
"""

import os
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────
INPUT_DIR  = "123"          # папка с файлами — поменяй если нужно
OUTPUT     = "MERGED_FINAL.xlsx"
MIN_SCORE  = 8            # минимальная оценка

# категории которые оставляем
VALID_CATS = {
    "pms", "rms", "ota", "channel manager", "booking engine",
    "crs", "vacation rental"
}

# мусорные домены которые убираем даже если score высокий
JUNK_DOMAINS = {
    "booking.com", "expedia.com", "hotels.com", "agoda.com",
    "airbnb.com", "tripadvisor.com", "kayak.com", "priceline.com",
    "trivago.com", "hostelworld.com", "vrbo.com",
    "capterra.com", "g2.com", "softwareadvice.com", "getapp.com",
    "hoteltechreport.com", "thehotelgm.com",
    "wikipedia.org", "youtube.com", "facebook.com", "twitter.com",
    "instagram.com", "linkedin.com", "reddit.com",
    "sunweb.be", "sunweb.nl", "freelancer.com",
    "investor.wyn", "bestprice.vn", "unihotels.ru",
}

def get_domain(url):
    if not url: return ""
    url = str(url)
    m = re.search(r'https?://(?:www\.)?([^/]+)', url)
    return m.group(1).lower() if m else ""

def is_junk(domain):
    if not domain: return True
    for j in JUNK_DOMAINS:
        if domain == j or domain.endswith("." + j): return True
    return False

def normalize_category(cat):
    if not cat: return "Unknown"
    cat = str(cat).strip()
    cl = cat.lower()
    if "property management" in cl or cl == "pms": return "PMS"
    if "channel manager" in cl: return "Channel Manager"
    if "booking engine" in cl: return "Booking Engine"
    if "reservation management" in cl or cl == "rms": return "RMS"
    if "central reservation" in cl or cl == "crs": return "CRS"
    if "online travel" in cl or cl == "ota": return "OTA"
    if "vacation rental" in cl: return "Vacation Rental"
    if "marketplace" in cl or "integrat" in cl: return "Marketplace"
    return cat

# ─── ЧИТАЛКИ ──────────────────────────────────────────────

def read_batch_xlsx(filepath):
    """batch_*.xlsx — ищем листы с высокими оценками"""
    rows = []
    try:
        wb = load_workbook(filepath, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h else "" for h in row]
                    continue
                d = dict(zip(headers, row))
                score = d.get("Score") or d.get("score")
                try: score = float(score)
                except: continue
                if score < MIN_SCORE: continue
                url = d.get("URL") or d.get("url") or ""
                domain = d.get("Domain") or d.get("domain") or get_domain(url)
                cat = normalize_category(d.get("Category") or d.get("category") or "")
                title = d.get("Title") or d.get("title") or ""
                desc = d.get("Description") or d.get("snippet") or d.get("reasoning") or ""
                rows.append({
                    "domain": str(domain).lower().strip(),
                    "url": str(url).strip(),
                    "category": cat,
                    "score": score,
                    "title": str(title)[:200],
                    "description": str(desc)[:300],
                    "source": Path(filepath).name,
                })
    except Exception as e:
        print(f"  [!] Ошибка {filepath}: {e}")
    return rows


def read_validated_xlsx(filepath):
    """validated_vendors.xlsx"""
    rows = []
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else "" for h in row]
                continue
            d = dict(zip(headers, row))
            score = d.get("Score") or d.get("score")
            try: score = float(score)
            except: continue
            if score < MIN_SCORE: continue
            verdict = str(d.get("Verdict") or "").upper()
            if verdict in ("NO", "SKIP"): continue
            url = d.get("URL") or d.get("url") or ""
            domain = d.get("Domain") or d.get("domain") or get_domain(url)
            cat = normalize_category(d.get("Category") or d.get("Sheet") or "")
            rows.append({
                "domain": str(domain).lower().strip(),
                "url": str(url).strip(),
                "category": cat,
                "score": score,
                "title": "",
                "description": str(d.get("Reason") or "")[:300],
                "source": Path(filepath).name,
            })
    except Exception as e:
        print(f"  [!] Ошибка {filepath}: {e}")
    return rows


def read_csv_file(filepath):
    """classified_sites / results CSV"""
    rows = []
    try:
        df = pd.read_csv(filepath)
        df.columns = [c.lower().strip() for c in df.columns]
        for _, r in df.iterrows():
            score = r.get("score", 0)
            try: score = float(score)
            except: continue
            if score < MIN_SCORE: continue
            url = str(r.get("url", "")).strip()
            domain = str(r.get("domain", get_domain(url))).strip().lower()
            cat = normalize_category(r.get("category", ""))
            rows.append({
                "domain": domain,
                "url": url,
                "category": cat,
                "score": score,
                "title": str(r.get("title", ""))[:200],
                "description": str(r.get("reasoning", r.get("description", "")))[:300],
                "source": Path(filepath).name,
            })
    except Exception as e:
        print(f"  [!] Ошибка {filepath}: {e}")
    return rows


# ─── MAIN ─────────────────────────────────────────────────
def main():
    input_path = Path(INPUT_DIR)
    all_rows = []

    files = sorted(input_path.glob("*.xlsx")) + sorted(input_path.glob("*.csv"))
    print(f"Найдено файлов: {len(files)}")

    for f in files:
        if f.name == OUTPUT: continue
        print(f"  Читаю: {f.name}")

        if f.suffix == ".csv":
            rows = read_csv_file(f)
        elif "validated" in f.name.lower():
            rows = read_validated_xlsx(f)
        elif "batch" in f.name.lower() or "scored" in f.name.lower():
            rows = read_batch_xlsx(f)
        else:
            rows = read_batch_xlsx(f)  # пробуем как batch

        print(f"    → {len(rows)} записей со score≥{MIN_SCORE}")
        all_rows.extend(rows)

    print(f"\nВсего записей до дедупликации: {len(all_rows)}")

    # фильтруем мусор
    all_rows = [r for r in all_rows if not is_junk(r["domain"])]
    all_rows = [r for r in all_rows if r["domain"]]
    keep_cats = {"PMS","RMS","CRS","OTA","Channel Manager","Booking Engine","Vacation Rental","Marketplace"}
    all_rows = [r for r in all_rows if r["category"] in keep_cats]

    # дедупликация по домену — оставляем с максимальным score
    domain_best = {}
    for r in all_rows:
        d = r["domain"]
        if d not in domain_best or r["score"] > domain_best[d]["score"]:
            domain_best[d] = r

    final = sorted(domain_best.values(), key=lambda x: (-x["score"], x["category"], x["domain"]))
    print(f"Уникальных доменов после дедупликации: {len(final)}")

    # статистика по категориям
    from collections import Counter
    cats = Counter(r["category"] for r in final)
    print("\nПо категориям:")
    for cat, cnt in cats.most_common():
        print(f"  {cat}: {cnt}")

    # ─── СОХРАНЯЕМ В EXCEL ────────────────────────────────
    df = pd.DataFrame(final, columns=["domain","url","category","score","title","description","source"])
    df = df.sort_values(["category","score"], ascending=[True, False])

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        # главный лист — все
        df.to_excel(writer, sheet_name="ALL", index=False)

        # листы по категориям
        for cat in df["category"].unique():
            sub = df[df["category"] == cat].copy()
            # убираем недопустимые символы Excel
            sheet_name = re.sub(r'[\\/*?\[\]:]', '_', str(cat))[:31]
            sub.to_excel(writer, sheet_name=sheet_name, index=False)

        # форматирование
        wb = writer.book
        header_fill = PatternFill("solid", start_color="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        green_fill  = PatternFill("solid", start_color="C6EFCE")
        yellow_fill = PatternFill("solid", start_color="FFEB9C")

        for sheet in wb.worksheets:
            # заголовки
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # ширины колонок
            widths = {"domain": 30, "url": 55, "category": 18,
                      "score": 7, "title": 35, "description": 60, "source": 25}
            for i, col in enumerate(df.columns, 1):
                sheet.column_dimensions[get_column_letter(i)].width = widths.get(col, 15)

            # цвет по score
            for row in sheet.iter_rows(min_row=2):
                try:
                    score_val = float(row[3].value or 0)
                    fill = green_fill if score_val >= 9 else yellow_fill
                    for cell in row:
                        cell.fill = fill
                        cell.font = Font(name="Arial", size=9)
                except: pass

    print(f"\n✅ Сохранено → {OUTPUT}")
    print(f"   Итого: {len(final)} уникальных вендоров")


if __name__ == "__main__":
    main()